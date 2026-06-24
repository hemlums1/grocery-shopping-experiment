#!/usr/bin/env python3
"""Builds an analysis workbook (reports/analysis-report-{date}.xlsx) from live
respondent data, with pivot-style tables and charts answering the experiment's
core research questions. Pulls tidy data from /admin/export.csv (the system of
record) and cross-references src/catalog.js for the "what else was available"
baseline that premium-paid and price-rank metrics need.

Run: python3 scripts/build_analysis_report.py [--base-url URL] [--out PATH]
"""

import argparse
import csv
import io
import json
import os
import subprocess
import urllib.request
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="2F5233")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=13)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
EXAMPLE_FILL = PatternFill("solid", start_color="FCE4D6")
EXAMPLE_FONT = Font(name=FONT, bold=True, color="9C3B00", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DIMENSIONS = [
    ("Brand Tier", "brand_tier"),
    ("Organic", "organic"),
    ("Sourcing Practice", "sourcing_practice"),
    ("Nutri-Score", "nutri_score"),
]


def load_env():
    env = {}
    with open(os.path.join(ROOT, ".env")) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def fetch_raw_rows(base_url, export_key):
    url = f"{base_url}/admin/export.csv?key={export_key}"
    with urllib.request.urlopen(url) as resp:
        text = resp.read().decode("utf-8")
    return list(csv.DictReader(io.StringIO(text)))


def fetch_catalog():
    result = subprocess.run(
        ["node", "-e", "console.log(JSON.stringify(require('./src/catalog').CATALOG))"],
        cwd=ROOT, capture_output=True, text=True, check=True,
    )
    return json.loads(result.stdout)


def build_baselines(catalog):
    by_type = defaultdict(list)
    for item in catalog:
        by_type[item["productType"]].append(item)
    min_price_by_type = {}
    rank_by_id = {}
    for product_type, items in by_type.items():
        ordered = sorted(items, key=lambda i: i["priceCents"])
        min_price_by_type[product_type] = ordered[0]["priceCents"]
        for rank, item in enumerate(ordered, start=1):
            rank_by_id[item["id"]] = rank
    return min_price_by_type, rank_by_id


def enrich(rows, min_price_by_type, rank_by_id):
    """Filters to submitted sessions with a non-empty basket and attaches
    premium_cents (vs. the cheapest available variant in the same product
    type) and rank (1=cheapest variant in its product type) to each row.
    Sessions with an empty basket are intentionally dropped here — see
    submitted_session_bases(), which is the denominator-safe session list."""
    enriched = []
    for r in rows:
        if r["status"] != "submitted" or not r["item_id"]:
            continue
        qty = int(r["quantity"])
        unit_cents = round(float(r["unit_price_usd"]) * 100)
        product_type = r["product_type"]
        out = dict(r)
        out["quantity"] = qty
        out["unit_cents"] = unit_cents
        out["premium_cents"] = unit_cents - min_price_by_type.get(product_type, unit_cents)
        out["rank"] = rank_by_id.get(r["item_id"])
        enriched.append(out)
    return enriched


def submitted_session_bases(raw_rows):
    """One entry per submitted session (session_id -> condition/budget),
    including sessions whose basket was empty at submission — the raw CSV
    guarantees exactly one row per session regardless of basket size, so this
    is the correct denominator for "how many people submitted," unlike the
    line-item-filtered rows from enrich()."""
    bases = {}
    for r in raw_rows:
        if r["status"] == "submitted" and r["session_id"] not in bases:
            bases[r["session_id"]] = {"condition": r["condition"], "budget_usd": float(r["budget_usd"])}
    return bases


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def aggregate_respondent_summary(rows, bases):
    by_session = defaultdict(list)
    for r in rows:
        by_session[r["session_id"]].append(r)

    out = []
    for session_id, base in bases.items():
        items = by_session.get(session_id, [])
        total_qty = sum(i["quantity"] for i in items)
        total_cents = sum(i["unit_cents"] * i["quantity"] for i in items)
        organic_qty = sum(i["quantity"] for i in items if i["organic"] == "Yes")
        brand_qty = sum(i["quantity"] for i in items if i["brand_tier"] == "Brand")
        rank_weighted = sum(i["rank"] * i["quantity"] for i in items if i["rank"])
        budget_cents = round(base["budget_usd"] * 100)
        out.append({
            "session_id": session_id,
            "condition": base["condition"],
            "budget_usd": base["budget_usd"],
            "item_count": total_qty,
            "distinct_categories": len({i["category"] for i in items}),
            "total_spent_usd": total_cents / 100,
            "pct_budget_used": (total_cents / budget_cents) if budget_cents else 0,
            "pct_organic_units": (organic_qty / total_qty) if total_qty else 0,
            "pct_brand_name_units": (brand_qty / total_qty) if total_qty else 0,
            "avg_price_rank": (rank_weighted / total_qty) if total_qty else 0,
        })
    out.sort(key=lambda r: r["session_id"])
    return out


def aggregate_attribute_preference(rows):
    total_units = sum(r["quantity"] for r in rows)
    table = []
    for label, field in DIMENSIONS:
        groups = defaultdict(lambda: [0, 0])  # level -> [units, premium_cents*qty]
        for r in rows:
            g = groups[r[field]]
            g[0] += r["quantity"]
            g[1] += r["premium_cents"] * r["quantity"]
        for level, (units, premium_total) in groups.items():
            table.append({
                "dimension": label,
                "level": level,
                "units": units,
                "pct_of_units": (units / total_units) if total_units else 0,
                "avg_premium_usd": (premium_total / units / 100) if units else 0,
            })
    table.sort(key=lambda r: -r["avg_premium_usd"])
    return table


def aggregate_budget_attribute(rows):
    totals_by_condition = defaultdict(int)
    for r in rows:
        totals_by_condition[r["condition"]] += r["quantity"]

    table = []
    for label, field in DIMENSIONS:
        groups = defaultdict(lambda: [0, 0])  # (level, condition) -> [units, premium_cents*qty]
        for r in rows:
            key = (r[field], r["condition"])
            g = groups[key]
            g[0] += r["quantity"]
            g[1] += r["premium_cents"] * r["quantity"]
        for (level, condition), (units, premium_total) in groups.items():
            denom = totals_by_condition[condition]
            table.append({
                "dimension": label,
                "level": level,
                "condition": condition,
                "pct_of_units": (units / denom) if denom else 0,
                "avg_premium_usd": (premium_total / units / 100) if units else 0,
            })
    table.sort(key=lambda r: (r["dimension"], r["level"], r["condition"]))
    return table


def aggregate_category_popularity(rows):
    total_units = sum(r["quantity"] for r in rows)
    total_cents = sum(r["quantity"] * r["unit_cents"] for r in rows)

    def grouped(field):
        groups = defaultdict(lambda: [0, 0])
        for r in rows:
            g = groups[r[field]]
            g[0] += r["quantity"]
            g[1] += r["quantity"] * r["unit_cents"]
        result = [
            {
                "name": name,
                "units": units,
                "pct_of_units": (units / total_units) if total_units else 0,
                "spend_usd": cents / 100,
                "pct_of_spend": (cents / total_cents) if total_cents else 0,
            }
            for name, (units, cents) in groups.items()
        ]
        result.sort(key=lambda r: -r["units"])
        return result

    return grouped("category"), grouped("product_type")


def aggregate_trade_down(rows):
    totals_by_condition = defaultdict(int)
    for r in rows:
        totals_by_condition[r["condition"]] += r["quantity"]

    groups = defaultdict(int)  # (condition, rank) -> units
    for r in rows:
        groups[(r["condition"], r["rank"])] += r["quantity"]

    conditions = sorted(totals_by_condition.keys(), key=lambda c: {"low": 0, "medium": 1, "high": 2}.get(c, 99))
    ranks = sorted({rank for _, rank in groups.keys() if rank})
    table = []
    for condition in conditions:
        denom = totals_by_condition[condition]
        row = {"condition": condition}
        for rank in ranks:
            units = groups.get((condition, rank), 0)
            row[f"rank_{rank}_pct"] = (units / denom) if denom else 0
        table.append(row)
    return table, ranks


# ---------------------------------------------------------------------------
# Sheet writing helpers
# ---------------------------------------------------------------------------

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def write_table(ws, start_row, start_col, headers, data_rows, percent_cols=(), currency_cols=()):
    """Writes a header row + data rows starting at (start_row, start_col).
    percent_cols/currency_cols are 0-indexed column positions within headers
    to apply number formatting to. Returns the row just after the table."""
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + j, value=h)
    style_header_row(ws, start_row, len(headers))

    for i, data_row in enumerate(data_rows):
        r = start_row + 1 + i
        for j, value in enumerate(data_row):
            cell = ws.cell(row=r, column=start_col + j, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            if j in percent_cols:
                cell.number_format = "0.0%"
            elif j in currency_cols:
                cell.number_format = '$#,##0.00'
    return start_row + 1 + len(data_rows)


def autosize_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_bar_chart(ws, anchor, title, data_ref, cats_ref, y_title, bar_dir="col", grouping="clustered", overlap=None):
    chart = BarChart()
    chart.type = bar_dir
    chart.grouping = grouping
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = None
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 18
    if overlap is not None:
        chart.overlap = overlap
    ws.add_chart(chart, anchor)
    return chart


def sheet_title(ws, text):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

def write_raw_data_sheet(wb, raw_rows, columns):
    ws = wb.create_sheet("Raw Data")
    for j, h in enumerate(columns, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(columns))
    for i, r in enumerate(raw_rows, start=2):
        for j, h in enumerate(columns, start=1):
            ws.cell(row=i, column=j, value=r.get(h, "")).font = BODY_FONT
    autosize_columns(ws, [16] + [12] * (len(columns) - 1))
    ws.freeze_panes = "A2"


def write_respondent_summary_sheet(wb, summary):
    ws = wb.create_sheet("Respondent Summary")
    headers = ["Session ID", "Condition", "Budget ($)", "Items Purchased", "Distinct Categories",
               "Total Spent ($)", "% Budget Used", "% Organic Units", "% Brand-Name Units", "Avg Price Rank Chosen"]
    rows = [[
        s["session_id"], s["condition"], s["budget_usd"], s["item_count"], s["distinct_categories"],
        s["total_spent_usd"], s["pct_budget_used"], s["pct_organic_units"], s["pct_brand_name_units"], s["avg_price_rank"],
    ] for s in summary]
    write_table(ws, 1, 1, headers, rows, percent_cols=(6, 7, 8), currency_cols=(2, 5))
    autosize_columns(ws, [38, 10, 11, 14, 16, 13, 12, 14, 16, 18])
    ws.freeze_panes = "A2"


def write_attribute_preference_sheet(wb, table):
    ws = wb.create_sheet("Attribute Preference")
    headers = ["Dimension", "Level", "Units Purchased", "% of Units", "Avg Premium Paid ($)"]
    rows = [[r["dimension"], r["level"], r["units"], r["pct_of_units"], r["avg_premium_usd"]] for r in table]
    end_row = write_table(ws, 1, 1, headers, rows, percent_cols=(3,), currency_cols=(4,))
    autosize_columns(ws, [20, 28, 16, 12, 20])

    # Two charts for legibility: the compact dimensions, and Sourcing Practice
    # on its own since it has many more levels than the others.
    core = [r for r in table if r["dimension"] != "Sourcing Practice"]
    sourcing = [r for r in table if r["dimension"] == "Sourcing Practice"]
    core.sort(key=lambda r: -r["avg_premium_usd"])
    sourcing.sort(key=lambda r: -r["avg_premium_usd"])

    chart_row = end_row + 2
    write_table(ws, chart_row, 1, ["Level", "Avg Premium Paid ($)"],
                [[f"{r['dimension']}: {r['level']}", r["avg_premium_usd"]] for r in core])
    data_ref = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(core))
    cats_ref = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(core))
    add_bar_chart(ws, f"D{chart_row}", "Avg Premium Paid — Brand / Organic / Nutri-Score", data_ref, cats_ref,
                  "Avg Premium Paid ($)", bar_dir="bar")

    chart_row2 = chart_row + len(core) + 2
    write_table(ws, chart_row2, 1, ["Sourcing Practice", "Avg Premium Paid ($)"],
                [[r["level"], r["avg_premium_usd"]] for r in sourcing])
    data_ref2 = Reference(ws, min_col=2, min_row=chart_row2, max_row=chart_row2 + len(sourcing))
    cats_ref2 = Reference(ws, min_col=1, min_row=chart_row2 + 1, max_row=chart_row2 + len(sourcing))
    add_bar_chart(ws, f"D{chart_row2}", "Avg Premium Paid — Sourcing Practice", data_ref2, cats_ref2,
                  "Avg Premium Paid ($)", bar_dir="bar")


def write_budget_attribute_sheet(wb, table):
    ws = wb.create_sheet("Budget x Attribute")
    headers = ["Dimension", "Level", "Condition", "% of Units (within condition)", "Avg Premium Paid ($)"]
    rows = [[r["dimension"], r["level"], r["condition"], r["pct_of_units"], r["avg_premium_usd"]] for r in table]
    end_row = write_table(ws, 1, 1, headers, rows, percent_cols=(3,), currency_cols=(4,))
    autosize_columns(ws, [20, 28, 11, 26, 20])

    chart_row = end_row + 2
    conditions = ["low", "medium", "high"]
    for label, field in DIMENSIONS:
        levels = sorted({r["level"] for r in table if r["dimension"] == label})
        by_level = {lvl: {} for lvl in levels}
        for r in table:
            if r["dimension"] == label:
                by_level[r["level"]][r["condition"]] = r["pct_of_units"]

        headers2 = ["Level"] + [c.capitalize() for c in conditions]
        rows2 = [[lvl] + [by_level[lvl].get(c, 0) for c in conditions] for lvl in levels]
        write_table(ws, chart_row, 1, headers2, rows2, percent_cols=(1, 2, 3))
        data_ref = Reference(ws, min_col=2, max_col=4, min_row=chart_row, max_row=chart_row + len(levels))
        cats_ref = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(levels))
        add_bar_chart(ws, f"G{chart_row}", f"% of Units by Budget Condition — {label}", data_ref, cats_ref, "% of Units")
        chart_row += max(len(levels) + 11, 17)  # charts are ~15 rows tall; ensure no overlap when a dimension has few levels


def write_category_popularity_sheet(wb, by_category, by_product_type):
    ws = wb.create_sheet("Category Popularity")
    sheet_title(ws, "By Category")
    headers = ["Category", "Units Purchased", "% of Units", "Spend ($)", "% of Spend"]
    rows = [[c["name"], c["units"], c["pct_of_units"], c["spend_usd"], c["pct_of_spend"]] for c in by_category]
    end_row = write_table(ws, 2, 1, headers, rows, percent_cols=(2, 4), currency_cols=(3,))
    autosize_columns(ws, [26, 16, 12, 12, 12])

    data_ref = Reference(ws, min_col=2, min_row=2, max_row=end_row - 1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=end_row - 1)
    add_bar_chart(ws, "G2", "Units Purchased by Category", data_ref, cats_ref, "Units")

    data_ref2 = Reference(ws, min_col=4, min_row=2, max_row=end_row - 1)
    add_bar_chart(ws, "G18", "Spend ($) by Category", data_ref2, cats_ref, "Spend ($)")

    type_row = end_row + 2
    ws.cell(row=type_row, column=1, value="By Product Type (detail, no chart)").font = TITLE_FONT
    headers_t = ["Product Type", "Units Purchased", "% of Units", "Spend ($)", "% of Spend"]
    rows_t = [[t["name"], t["units"], t["pct_of_units"], t["spend_usd"], t["pct_of_spend"]] for t in by_product_type]
    write_table(ws, type_row + 1, 1, headers_t, rows_t, percent_cols=(2, 4), currency_cols=(3,))


def write_trade_down_sheet(wb, table, ranks):
    ws = wb.create_sheet("Trade-Down Analysis")
    headers = ["Condition"] + [f"Rank {r} ({'cheapest' if r == ranks[0] else 'priciest' if r == ranks[-1] else 'mid'})" for r in ranks]
    rows = [[t["condition"].capitalize()] + [t[f"rank_{r}_pct"] for r in ranks] for t in table]
    end_row = write_table(ws, 1, 1, headers, rows, percent_cols=tuple(range(1, len(ranks) + 1)))
    autosize_columns(ws, [12] + [18] * len(ranks))

    data_ref = Reference(ws, min_col=2, max_col=1 + len(ranks), min_row=1, max_row=end_row - 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=end_row - 1)
    add_bar_chart(ws, f"A{end_row + 2}", "Price-Rank Chosen by Budget Condition (% within condition)",
                  data_ref, cats_ref, "% of Units", grouping="stacked", overlap=100)

    note_row = end_row + 20
    ws.cell(row=note_row, column=1,
            value="Rank 1 = cheapest variant available in that product type; higher ranks = pricier variants.").font = NOTE_FONT


def write_example_sheet(wb):
    ws = wb.create_sheet("Example (Fake Data)", 0)
    ws["A1"] = "EXAMPLE SHEET — fabricated numbers, for chart-format preview only. Not real respondent data."
    ws["A1"].font = EXAMPLE_FONT
    ws["A1"].fill = EXAMPLE_FILL
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = EXAMPLE_FILL
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 22

    # Attribute Preference example
    sheet_title(ws, "")
    ws["A3"] = "Attribute Preference (example)"
    ws["A3"].font = TITLE_FONT
    headers = ["Dimension", "Level", "% of Units", "Avg Premium Paid ($)"]
    rows = [
        ["Brand Tier", "Off-Brand", 0.41, 0.00],
        ["Brand Tier", "Brand", 0.59, 0.87],
        ["Organic", "No", 0.63, 0.00],
        ["Organic", "Yes", 0.37, 1.42],
        ["Sourcing", "Free-Range", 0.22, 1.15],
        ["Sourcing", "Factory-Farmed", 0.18, 0.00],
        ["Nutri-Score", "A", 0.35, 0.30],
        ["Nutri-Score", "E", 0.12, 0.05],
    ]
    end_row = write_table(ws, 4, 1, headers, rows, percent_cols=(2,), currency_cols=(3,))
    data_ref = Reference(ws, min_col=3, min_row=4, max_row=end_row - 1)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=end_row - 1)
    add_bar_chart(ws, "F4", "Example: % of Units by Attribute Level", data_ref, cats_ref, "% of Units", bar_dir="bar")

    # Budget x Attribute example
    row2 = end_row + 16
    ws.cell(row=row2, column=1, value="Budget x Attribute (example)").font = TITLE_FONT
    headers2 = ["Condition", "% Organic Units"]
    rows2 = [["Low", 0.18], ["Medium", 0.34], ["High", 0.52]]
    end_row2 = write_table(ws, row2 + 1, 1, headers2, rows2, percent_cols=(1,))
    data_ref2 = Reference(ws, min_col=2, min_row=row2 + 1, max_row=end_row2 - 1)
    cats_ref2 = Reference(ws, min_col=1, min_row=row2 + 2, max_row=end_row2 - 1)
    add_bar_chart(ws, f"F{row2}", "Example: % Organic by Budget Condition", data_ref2, cats_ref2, "% of Units")

    # Category Popularity example
    row3 = end_row2 + 16
    ws.cell(row=row3, column=1, value="Category Popularity (example)").font = TITLE_FONT
    headers3 = ["Category", "Units", "Spend ($)"]
    rows3 = [
        ["Produce", 120, 310], ["Pantry & Condiments", 95, 260], ["Dairy & Eggs", 80, 290],
        ["Snacks & Beverages", 70, 180], ["Meat & Seafood", 40, 310],
    ]
    end_row3 = write_table(ws, row3 + 1, 1, headers3, rows3, currency_cols=(2,))
    data_ref3 = Reference(ws, min_col=2, min_row=row3 + 1, max_row=end_row3 - 1)
    cats_ref3 = Reference(ws, min_col=1, min_row=row3 + 2, max_row=end_row3 - 1)
    add_bar_chart(ws, f"F{row3}", "Example: Units Purchased by Category", data_ref3, cats_ref3, "Units")

    # Trade-Down example
    row4 = end_row3 + 16
    ws.cell(row=row4, column=1, value="Trade-Down Analysis (example)").font = TITLE_FONT
    headers4 = ["Condition", "Rank 1 (cheapest)", "Rank 2", "Rank 3", "Rank 4 (priciest)"]
    rows4 = [
        ["Low", 0.55, 0.25, 0.15, 0.05],
        ["Medium", 0.35, 0.30, 0.20, 0.15],
        ["High", 0.20, 0.25, 0.25, 0.30],
    ]
    end_row4 = write_table(ws, row4 + 1, 1, headers4, rows4, percent_cols=(1, 2, 3, 4))
    data_ref4 = Reference(ws, min_col=2, max_col=5, min_row=row4 + 1, max_row=end_row4 - 1)
    cats_ref4 = Reference(ws, min_col=1, min_row=row4 + 2, max_row=end_row4 - 1)
    add_bar_chart(ws, f"F{row4}", "Example: Price-Rank Chosen by Condition", data_ref4, cats_ref4, "% of Units",
                  grouping="stacked", overlap=100)

    autosize_columns(ws, [22, 22, 14, 18])


# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", default="http://localhost:3000")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    env = load_env()
    export_key = env.get("EXPORT_KEY")
    if not export_key:
        raise SystemExit("EXPORT_KEY not found in .env")

    print(f"Fetching raw export from {args.base_url} ...")
    raw_rows = fetch_raw_rows(args.base_url, export_key)
    columns = list(raw_rows[0].keys()) if raw_rows else []
    print(f"  {len(raw_rows)} raw rows")

    print("Loading catalog for price-rank baselines ...")
    catalog = fetch_catalog()
    min_price_by_type, rank_by_id = build_baselines(catalog)

    rows = enrich(raw_rows, min_price_by_type, rank_by_id)
    bases = submitted_session_bases(raw_rows)
    print(f"  {len(rows)} submitted line items after filtering, {len(bases)} submitted sessions")

    wb = Workbook()
    wb.remove(wb.active)

    write_example_sheet(wb)
    write_raw_data_sheet(wb, raw_rows, columns)
    write_respondent_summary_sheet(wb, aggregate_respondent_summary(rows, bases))
    write_attribute_preference_sheet(wb, aggregate_attribute_preference(rows))
    write_budget_attribute_sheet(wb, aggregate_budget_attribute(rows))
    by_category, by_product_type = aggregate_category_popularity(rows)
    write_category_popularity_sheet(wb, by_category, by_product_type)
    trade_down_table, ranks = aggregate_trade_down(rows)
    write_trade_down_sheet(wb, trade_down_table, ranks)

    out_path = args.out or os.path.join(ROOT, "reports", f"analysis-report-{date.today().isoformat()}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
