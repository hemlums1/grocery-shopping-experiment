from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="2F5233")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
NUTRI_COLORS = {
    "A": "038141", "B": "85BB2F", "C": "FECB02", "D": "EF7D00", "E": "E63E11",
}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (item_id, category, product_type, item_name, brand, brand_tier, organic, sourcing_practice, country, nutri_score, price)
ROWS = [
    # Produce
    ("apples-sb-conv","Produce","Apples","EveryDay Basics Apples (2 lb bag)","EveryDay Basics","Off-Brand","No","Conventional","USA (Washington)","A",2.99),
    ("apples-brand-conv","Produce","Apples","Sunrise Farms Apples (2 lb bag)","Sunrise Farms","Brand","No","Conventional","USA (Washington)","A",3.79),
    ("apples-sb-organic","Produce","Apples","EveryDay Basics Organic Apples (2 lb bag)","EveryDay Basics","Off-Brand","Yes","Organic Farming","USA (Washington)","A",4.49),
    ("apples-brand-organic","Produce","Apples","Sunrise Farms Organic Apples (2 lb bag)","Sunrise Farms","Brand","Yes","Organic Farming","USA (Washington)","A",5.99),
    ("bananas-sb-std","Produce","Bananas","EveryDay Basics Bananas (per lb)","EveryDay Basics","Off-Brand","No","Standard Sourcing","Ecuador","A",0.59),
    ("bananas-brand-std","Produce","Bananas","Sunrise Farms Bananas (per lb)","Sunrise Farms","Brand","No","Standard Sourcing","Ecuador","A",0.79),
    ("bananas-sb-ft","Produce","Bananas","EveryDay Basics Fair Trade Bananas (per lb)","EveryDay Basics","Off-Brand","No","Fair-Trade Certified","Ecuador","A",0.99),
    ("bananas-brand-ft-org","Produce","Bananas","Sunrise Farms Fair Trade Organic Bananas (per lb)","Sunrise Farms","Brand","Yes","Fair-Trade Certified","Ecuador","A",1.49),
    ("tomatoes-sb-imp","Produce","Tomatoes","EveryDay Basics Tomatoes (per lb)","EveryDay Basics","Off-Brand","No","Imported","Mexico","A",1.49),
    ("tomatoes-brand-imp","Produce","Tomatoes","Sunrise Farms Tomatoes (per lb)","Sunrise Farms","Brand","No","Imported","Mexico","A",1.99),
    ("tomatoes-sb-local-org","Produce","Tomatoes","EveryDay Basics Local Organic Tomatoes (per lb)","EveryDay Basics","Off-Brand","Yes","Locally Grown","USA (California)","A",2.99),
    ("tomatoes-brand-local-org","Produce","Tomatoes","Sunrise Farms Local Organic Tomatoes (per lb)","Sunrise Farms","Brand","Yes","Locally Grown","USA (California)","A",3.79),
    ("spinach-sb-conv","Produce","Spinach","EveryDay Basics Spinach (5 oz bag)","EveryDay Basics","Off-Brand","No","Conventional","USA (California)","A",2.29),
    ("spinach-brand-conv","Produce","Spinach","Sunrise Farms Spinach (5 oz bag)","Sunrise Farms","Brand","No","Conventional","USA (California)","A",2.99),
    ("spinach-sb-organic","Produce","Spinach","EveryDay Basics Organic Spinach (5 oz bag)","EveryDay Basics","Off-Brand","Yes","Organic Farming","USA (California)","A",3.79),
    ("spinach-brand-organic","Produce","Spinach","Sunrise Farms Organic Spinach (5 oz bag)","Sunrise Farms","Brand","Yes","Organic Farming","USA (California)","A",4.59),
    ("potatoes-sb-conv","Produce","Potatoes","EveryDay Basics Russet Potatoes (5 lb bag)","EveryDay Basics","Off-Brand","No","Conventional","USA (Idaho)","B",3.49),
    ("potatoes-brand-conv","Produce","Potatoes","Sunrise Farms Russet Potatoes (5 lb bag)","Sunrise Farms","Brand","No","Conventional","USA (Idaho)","B",4.29),
    ("potatoes-sb-organic","Produce","Potatoes","EveryDay Basics Organic Russet Potatoes (5 lb bag)","EveryDay Basics","Off-Brand","Yes","Organic Farming","USA (Idaho)","B",5.49),
    ("potatoes-brand-organic","Produce","Potatoes","Sunrise Farms Organic Russet Potatoes (5 lb bag)","Sunrise Farms","Brand","Yes","Organic Farming","USA (Idaho)","B",6.49),
    # Dairy & Eggs
    ("milk-sb-conv","Dairy & Eggs","Milk","EveryDay Basics Whole Milk (1 gal)","EveryDay Basics","Off-Brand","No","Conventional Dairy","USA (Wisconsin)","C",3.29),
    ("milk-brand-conv","Dairy & Eggs","Milk","Green Pasture Co. Whole Milk (1 gal)","Green Pasture Co.","Brand","No","Conventional Dairy","USA (Wisconsin)","C",4.49),
    ("milk-sb-organic","Dairy & Eggs","Milk","EveryDay Basics Organic Whole Milk (1 gal)","EveryDay Basics","Off-Brand","Yes","Pasture-Raised","USA (Wisconsin)","C",5.29),
    ("milk-brand-organic","Dairy & Eggs","Milk","Green Pasture Co. Organic Grass-Fed Whole Milk (1 gal)","Green Pasture Co.","Brand","Yes","Pasture-Raised/Grass-Fed","USA (Vermont)","C",6.99),
    ("eggs-sb-caged","Dairy & Eggs","Eggs","EveryDay Basics Eggs (dozen, large)","EveryDay Basics","Off-Brand","No","Caged","USA (Iowa)","B",2.49),
    ("eggs-brand-caged","Dairy & Eggs","Eggs","Heritage Range Eggs (dozen, large)","Heritage Range","Brand","No","Caged","USA (Iowa)","B",3.29),
    ("eggs-sb-freerange","Dairy & Eggs","Eggs","EveryDay Basics Free-Range Eggs (dozen, large)","EveryDay Basics","Off-Brand","No","Free-Range","USA (Iowa)","B",4.49),
    ("eggs-brand-pasture-org","Dairy & Eggs","Eggs","Heritage Range Pasture-Raised Organic Eggs (dozen, large)","Heritage Range","Brand","Yes","Pasture-Raised","USA (Iowa)","B",6.29),
    ("cheese-sb-conv","Dairy & Eggs","Cheddar Cheese","EveryDay Basics Cheddar Cheese (8 oz block)","EveryDay Basics","Off-Brand","No","Conventional Dairy","USA (Wisconsin)","D",3.49),
    ("cheese-brand-conv","Dairy & Eggs","Cheddar Cheese","Green Pasture Co. Cheddar Cheese (8 oz block)","Green Pasture Co.","Brand","No","Conventional Dairy","USA (Wisconsin)","D",4.99),
    ("cheese-sb-organic","Dairy & Eggs","Cheddar Cheese","EveryDay Basics Organic Cheddar Cheese (8 oz block)","EveryDay Basics","Off-Brand","Yes","Pasture-Raised","USA (Vermont)","D",5.99),
    ("cheese-brand-organic","Dairy & Eggs","Cheddar Cheese","Green Pasture Co. Organic Grass-Fed Cheddar Cheese (8 oz block)","Green Pasture Co.","Brand","Yes","Pasture-Raised/Grass-Fed","USA (Vermont)","D",7.49),
    ("yogurt-sb-plain","Dairy & Eggs","Yogurt","EveryDay Basics Plain Yogurt (32 oz)","EveryDay Basics","Off-Brand","No","Conventional Dairy","USA","B",3.29),
    ("yogurt-brand-plain","Dairy & Eggs","Yogurt","Green Pasture Co. Plain Yogurt (32 oz)","Green Pasture Co.","Brand","No","Conventional Dairy","USA","B",4.49),
    ("yogurt-sb-organic-plain","Dairy & Eggs","Yogurt","EveryDay Basics Organic Plain Yogurt (32 oz)","EveryDay Basics","Off-Brand","Yes","Pasture-Raised","USA","B",5.49),
    ("yogurt-brand-organic-plain","Dairy & Eggs","Yogurt","Green Pasture Co. Organic Plain Yogurt (32 oz)","Green Pasture Co.","Brand","Yes","Pasture-Raised","USA","B",6.99),
    ("butter-sb-conv","Dairy & Eggs","Butter","EveryDay Basics Butter (1 lb)","EveryDay Basics","Off-Brand","No","Conventional Dairy","USA (Wisconsin)","E",3.99),
    ("butter-brand-conv","Dairy & Eggs","Butter","Green Pasture Co. Butter (1 lb)","Green Pasture Co.","Brand","No","Conventional Dairy","USA (Wisconsin)","E",5.49),
    ("butter-sb-organic","Dairy & Eggs","Butter","EveryDay Basics Organic Butter (1 lb)","EveryDay Basics","Off-Brand","Yes","Pasture-Raised","USA (Vermont)","E",6.49),
    ("butter-brand-organic","Dairy & Eggs","Butter","Green Pasture Co. Organic Grass-Fed Butter (1 lb)","Green Pasture Co.","Brand","Yes","Pasture-Raised/Grass-Fed","New Zealand","E",7.99),
    # Meat & Seafood
    ("chicken-sb-factory","Meat & Seafood","Chicken Breast","EveryDay Basics Chicken Breast (1 lb)","EveryDay Basics","Off-Brand","No","Factory-Farmed","USA (Arkansas)","A",3.99),
    ("chicken-brand-factory","Meat & Seafood","Chicken Breast","Heritage Range Chicken Breast (1 lb)","Heritage Range","Brand","No","Factory-Farmed","USA (Arkansas)","A",4.99),
    ("chicken-sb-freerange","Meat & Seafood","Chicken Breast","EveryDay Basics Free-Range Chicken Breast (1 lb)","EveryDay Basics","Off-Brand","No","Free-Range","USA (Pennsylvania)","A",6.49),
    ("chicken-brand-freerange-org","Meat & Seafood","Chicken Breast","Heritage Range Organic Free-Range Chicken Breast (1 lb)","Heritage Range","Brand","Yes","Free-Range","USA (Pennsylvania)","A",8.99),
    ("beef-sb-factory","Meat & Seafood","Ground Beef","EveryDay Basics Ground Beef 80/20 (1 lb)","EveryDay Basics","Off-Brand","No","Factory-Farmed (Grain-Fed)","USA (Nebraska)","C",5.49),
    ("beef-brand-factory","Meat & Seafood","Ground Beef","Heritage Range Ground Beef 80/20 (1 lb)","Heritage Range","Brand","No","Factory-Farmed (Grain-Fed)","USA (Nebraska)","C",6.99),
    ("beef-sb-grassfed","Meat & Seafood","Ground Beef","EveryDay Basics Grass-Fed Ground Beef (1 lb)","EveryDay Basics","Off-Brand","No","Grass-Fed/Pasture-Raised","USA (Montana)","C",8.49),
    ("beef-brand-grassfed-org","Meat & Seafood","Ground Beef","Heritage Range Organic Grass-Fed Ground Beef (1 lb)","Heritage Range","Brand","Yes","Grass-Fed/Pasture-Raised","New Zealand","C",10.99),
    ("bacon-sb-factory","Meat & Seafood","Bacon","EveryDay Basics Bacon (12 oz)","EveryDay Basics","Off-Brand","No","Factory-Farmed","USA (Iowa)","E",4.99),
    ("bacon-brand-factory","Meat & Seafood","Bacon","Heritage Range Bacon (12 oz)","Heritage Range","Brand","No","Factory-Farmed","USA (Iowa)","E",6.49),
    ("bacon-sb-freerange","Meat & Seafood","Bacon","EveryDay Basics Uncured Free-Range Bacon (12 oz)","EveryDay Basics","Off-Brand","No","Free-Range","USA (Iowa)","E",7.99),
    ("bacon-brand-freerange-org","Meat & Seafood","Bacon","Heritage Range Organic Free-Range Bacon (12 oz)","Heritage Range","Brand","Yes","Free-Range","USA (Iowa)","E",9.99),
    ("salmon-sb-farmed","Meat & Seafood","Salmon Fillet","EveryDay Basics Salmon Fillet (1 lb)","EveryDay Basics","Off-Brand","No","Farmed","Chile","A",7.99),
    ("salmon-brand-farmed","Meat & Seafood","Salmon Fillet","Heritage Range Salmon Fillet (1 lb)","Heritage Range","Brand","No","Farmed","Norway","A",9.49),
    ("salmon-sb-wild","Meat & Seafood","Salmon Fillet","EveryDay Basics Wild-Caught Salmon Fillet (1 lb)","EveryDay Basics","Off-Brand","No","Wild-Caught","USA (Alaska)","A",11.99),
    ("salmon-brand-wild-msc","Meat & Seafood","Salmon Fillet","Heritage Range Wild-Caught Salmon Fillet (1 lb)","Heritage Range","Brand","No","Wild-Caught, MSC Certified","USA (Alaska)","A",14.99),
    ("tuna-sb-std","Meat & Seafood","Canned Tuna","EveryDay Basics Chunk Tuna in Water (5 oz)","EveryDay Basics","Off-Brand","No","Standard","Thailand","A",1.29),
    ("tuna-brand-std","Meat & Seafood","Canned Tuna","Pinnacle Foods Chunk Tuna in Water (5 oz)","Pinnacle Foods","Brand","No","Standard","Thailand","A",1.99),
    ("tuna-sb-wild","Meat & Seafood","Canned Tuna","EveryDay Basics Wild-Caught Tuna in Water (5 oz)","EveryDay Basics","Off-Brand","No","Wild-Caught, Dolphin-Safe","USA","A",2.49),
    ("tuna-brand-wild-oil","Meat & Seafood","Canned Tuna","Pinnacle Foods Wild-Caught Tuna in Olive Oil (5 oz)","Pinnacle Foods","Brand","No","Wild-Caught, Dolphin-Safe","Italy","B",3.49),
    # Bakery & Grains
    ("bread-sb-white","Bakery & Grains","Bread","EveryDay Basics White Bread (24 oz loaf)","EveryDay Basics","Off-Brand","No","Standard","USA","C",2.49),
    ("bread-brand-white","Bakery & Grains","Bread","Golden Harvest White Bread (24 oz loaf)","Golden Harvest","Brand","No","Standard","USA","C",3.49),
    ("bread-sb-wheat","Bakery & Grains","Bread","EveryDay Basics Whole Wheat Bread (24 oz loaf)","EveryDay Basics","Off-Brand","No","Standard","USA","A",2.99),
    ("bread-brand-wheat-org","Bakery & Grains","Bread","Golden Harvest Organic Whole Wheat Bread (24 oz loaf)","Golden Harvest","Brand","Yes","Standard","USA","A",4.99),
    ("pasta-sb-reg","Bakery & Grains","Pasta","EveryDay Basics Spaghetti (1 lb box)","EveryDay Basics","Off-Brand","No","Standard","Italy","B",1.29),
    ("pasta-brand-reg","Bakery & Grains","Pasta","Golden Harvest Spaghetti (1 lb box)","Golden Harvest","Brand","No","Standard","Italy","B",1.99),
    ("pasta-sb-wheat","Bakery & Grains","Pasta","EveryDay Basics Whole Wheat Spaghetti (1 lb box)","EveryDay Basics","Off-Brand","No","Standard","Italy","A",1.79),
    ("pasta-brand-wheat-org","Bakery & Grains","Pasta","Golden Harvest Organic Whole Wheat Spaghetti (1 lb box)","Golden Harvest","Brand","Yes","Standard","Italy","A",3.49),
    ("rice-sb-white","Bakery & Grains","Rice","EveryDay Basics White Rice (2 lb bag)","EveryDay Basics","Off-Brand","No","Standard","USA","B",2.99),
    ("rice-brand-white","Bakery & Grains","Rice","Golden Harvest White Rice (2 lb bag)","Golden Harvest","Brand","No","Standard","USA","B",3.99),
    ("rice-sb-brown","Bakery & Grains","Rice","EveryDay Basics Brown Rice (2 lb bag)","EveryDay Basics","Off-Brand","No","Standard","USA","A",3.49),
    ("rice-brand-brown-org","Bakery & Grains","Rice","Golden Harvest Organic Brown Rice (2 lb bag)","Golden Harvest","Brand","Yes","Standard","USA","A",5.49),
    ("cereal-sb-corn","Bakery & Grains","Cereal","EveryDay Basics Corn Flakes (14 oz box)","EveryDay Basics","Off-Brand","No","Standard","USA","C",2.99),
    ("cereal-brand-frosted","Bakery & Grains","Cereal","Pinnacle Foods Frosted Corn Flakes (14 oz box)","Pinnacle Foods","Brand","No","Standard","USA","D",4.49),
    ("cereal-sb-bran","Bakery & Grains","Cereal","EveryDay Basics Bran Flakes (14 oz box)","EveryDay Basics","Off-Brand","No","Standard","USA","A",3.49),
    ("cereal-brand-bran-org","Bakery & Grains","Cereal","Pinnacle Foods Organic Bran Flakes (14 oz box)","Pinnacle Foods","Brand","Yes","Standard","USA","A",5.99),
    # Pantry & Condiments
    ("oil-sb-reg","Pantry & Condiments","Olive Oil","EveryDay Basics Olive Oil (16.9 oz)","EveryDay Basics","Off-Brand","No","Standard","Spain","C",5.99),
    ("oil-brand-evoo","Pantry & Condiments","Olive Oil","Golden Harvest Extra Virgin Olive Oil (16.9 oz)","Golden Harvest","Brand","No","Standard","Spain","C",8.99),
    ("oil-sb-organic","Pantry & Condiments","Olive Oil","EveryDay Basics Organic Extra Virgin Olive Oil (16.9 oz)","EveryDay Basics","Off-Brand","Yes","Standard","Italy","C",10.99),
    ("oil-brand-organic-estate","Pantry & Condiments","Olive Oil","Golden Harvest Organic Estate-Grown Olive Oil (16.9 oz)","Golden Harvest","Brand","Yes","Sustainably Harvested","Italy","C",14.99),
    ("tomcan-sb-reg","Pantry & Condiments","Canned Tomatoes","EveryDay Basics Diced Tomatoes (28 oz can)","EveryDay Basics","Off-Brand","No","Standard","USA","A",1.49),
    ("tomcan-brand-reg","Pantry & Condiments","Canned Tomatoes","Golden Harvest Diced Tomatoes (28 oz can)","Golden Harvest","Brand","No","Standard","USA","A",2.29),
    ("tomcan-sb-organic","Pantry & Condiments","Canned Tomatoes","EveryDay Basics Organic Diced Tomatoes (28 oz can)","EveryDay Basics","Off-Brand","Yes","Standard","USA","A",2.99),
    ("tomcan-brand-organic-sm","Pantry & Condiments","Canned Tomatoes","Golden Harvest Organic San Marzano Diced Tomatoes (28 oz can)","Golden Harvest","Brand","Yes","Standard","Italy","A",4.49),
    ("pb-sb-reg","Pantry & Condiments","Peanut Butter","EveryDay Basics Peanut Butter (16 oz jar)","EveryDay Basics","Off-Brand","No","Standard","USA (Georgia)","D",2.99),
    ("pb-brand-reg","Pantry & Condiments","Peanut Butter","Pinnacle Foods Peanut Butter (16 oz jar)","Pinnacle Foods","Brand","No","Standard","USA (Georgia)","D",4.29),
    ("pb-sb-natural","Pantry & Condiments","Peanut Butter","EveryDay Basics Natural Peanut Butter, No Sugar Added (16 oz jar)","EveryDay Basics","Off-Brand","No","Standard","USA (Georgia)","C",4.49),
    ("pb-brand-natural-org","Pantry & Condiments","Peanut Butter","Pinnacle Foods Organic Natural Peanut Butter (16 oz jar)","Pinnacle Foods","Brand","Yes","Standard","USA (Georgia)","C",6.49),
    ("coffee-sb-std","Pantry & Condiments","Coffee","EveryDay Basics Ground Coffee (12 oz bag)","EveryDay Basics","Off-Brand","No","Standard Sourcing","Brazil","A",6.99),
    ("coffee-brand-std","Pantry & Condiments","Coffee","Pinnacle Foods Ground Coffee (12 oz bag)","Pinnacle Foods","Brand","No","Standard Sourcing","Brazil","A",8.99),
    ("coffee-sb-ft","Pantry & Condiments","Coffee","EveryDay Basics Fair Trade Ground Coffee (12 oz bag)","EveryDay Basics","Off-Brand","No","Fair-Trade Certified","Colombia","A",10.49),
    ("coffee-brand-ft-org","Pantry & Condiments","Coffee","Pinnacle Foods Fair Trade Organic Ground Coffee (12 oz bag)","Pinnacle Foods","Brand","Yes","Fair-Trade Certified","Ethiopia","A",13.99),
    # Snacks & Beverages
    ("choc-sb-milk","Snacks & Beverages","Chocolate Bar","EveryDay Basics Milk Chocolate Bar (3.5 oz)","EveryDay Basics","Off-Brand","No","Standard Sourcing","Cote d'Ivoire","E",1.99),
    ("choc-brand-milk","Snacks & Beverages","Chocolate Bar","Pinnacle Foods Milk Chocolate Bar (3.5 oz)","Pinnacle Foods","Brand","No","Standard Sourcing","Cote d'Ivoire","E",2.99),
    ("choc-sb-dark-ft","Snacks & Beverages","Chocolate Bar","EveryDay Basics Fair Trade Dark Chocolate Bar 70% (3.5 oz)","EveryDay Basics","Off-Brand","No","Fair-Trade Certified","Ghana","D",3.49),
    ("choc-brand-dark-ft-org","Snacks & Beverages","Chocolate Bar","Pinnacle Foods Fair Trade Organic Dark Chocolate Bar 70% (3.5 oz)","Pinnacle Foods","Brand","Yes","Fair-Trade Certified","Ecuador","D",4.99),
    ("chips-sb-reg","Snacks & Beverages","Potato Chips","EveryDay Basics Potato Chips (8 oz bag)","EveryDay Basics","Off-Brand","No","Standard","USA (Idaho)","E",2.49),
    ("chips-brand-reg","Snacks & Beverages","Potato Chips","Pinnacle Foods Potato Chips (8 oz bag)","Pinnacle Foods","Brand","No","Standard","USA (Idaho)","E",3.79),
    ("chips-sb-reducedfat","Snacks & Beverages","Potato Chips","EveryDay Basics Reduced Fat Potato Chips (8 oz bag)","EveryDay Basics","Off-Brand","No","Standard","USA (Idaho)","D",3.29),
    ("chips-brand-avocado-org","Snacks & Beverages","Potato Chips","Pinnacle Foods Organic Potato Chips, Avocado Oil (8 oz bag)","Pinnacle Foods","Brand","Yes","Standard","USA (Idaho)","D",4.99),
    ("oj-sb-conc","Snacks & Beverages","Orange Juice","EveryDay Basics Orange Juice, From Concentrate (64 oz)","EveryDay Basics","Off-Brand","No","Standard","USA (Florida)","C",2.99),
    ("oj-brand-conc","Snacks & Beverages","Orange Juice","Sunrise Farms Orange Juice, From Concentrate (64 oz)","Sunrise Farms","Brand","No","Standard","USA (Florida)","C",4.29),
    ("oj-sb-nfc","Snacks & Beverages","Orange Juice","EveryDay Basics Not-From-Concentrate Orange Juice (64 oz)","EveryDay Basics","Off-Brand","No","Standard","USA (Florida)","B",4.99),
    ("oj-brand-nfc-org","Snacks & Beverages","Orange Juice","Sunrise Farms Organic Not-From-Concentrate Orange Juice (64 oz)","Sunrise Farms","Brand","Yes","Standard","USA (Florida)","B",6.99),
    # Produce (additions)
    ("onions-sb-conv","Produce","Onions","EveryDay Basics Yellow Onions (3 lb bag)","EveryDay Basics","Off-Brand","No","Conventional","USA (Oregon)","A",2.49),
    ("onions-brand-conv","Produce","Onions","Sunrise Farms Yellow Onions (3 lb bag)","Sunrise Farms","Brand","No","Conventional","USA (Oregon)","A",3.29),
    ("onions-sb-organic","Produce","Onions","EveryDay Basics Organic Yellow Onions (3 lb bag)","EveryDay Basics","Off-Brand","Yes","Organic Farming","USA (Oregon)","A",3.99),
    ("onions-brand-organic","Produce","Onions","Sunrise Farms Organic Yellow Onions (3 lb bag)","Sunrise Farms","Brand","Yes","Organic Farming","USA (Oregon)","A",4.99),
    ("carrots-sb-conv","Produce","Carrots","EveryDay Basics Carrots (2 lb bag)","EveryDay Basics","Off-Brand","No","Conventional","USA (California)","A",1.99),
    ("carrots-brand-conv","Produce","Carrots","Sunrise Farms Carrots (2 lb bag)","Sunrise Farms","Brand","No","Conventional","USA (California)","A",2.69),
    ("carrots-sb-organic","Produce","Carrots","EveryDay Basics Organic Carrots (2 lb bag)","EveryDay Basics","Off-Brand","Yes","Organic Farming","USA (California)","A",3.49),
    ("carrots-brand-organic","Produce","Carrots","Sunrise Farms Organic Carrots (2 lb bag)","Sunrise Farms","Brand","Yes","Organic Farming","USA (California)","A",4.29),
    ("avocados-sb-imp","Produce","Avocados","EveryDay Basics Avocados (4 ct bag)","EveryDay Basics","Off-Brand","No","Imported","Mexico","B",3.99),
    ("avocados-brand-imp","Produce","Avocados","Sunrise Farms Avocados (4 ct bag)","Sunrise Farms","Brand","No","Imported","Mexico","B",4.99),
    ("avocados-sb-local-org","Produce","Avocados","EveryDay Basics Local Organic Avocados (4 ct bag)","EveryDay Basics","Off-Brand","Yes","Locally Grown","USA (California)","B",6.49),
    ("avocados-brand-local-org","Produce","Avocados","Sunrise Farms Local Organic Avocados (4 ct bag)","Sunrise Farms","Brand","Yes","Locally Grown","USA (California)","B",7.99),
    # Dairy & Eggs (additions)
    ("sourcream-sb-conv","Dairy & Eggs","Sour Cream","EveryDay Basics Sour Cream (16 oz)","EveryDay Basics","Off-Brand","No","Conventional Dairy","USA (Wisconsin)","D",1.99),
    ("sourcream-brand-conv","Dairy & Eggs","Sour Cream","Green Pasture Co. Sour Cream (16 oz)","Green Pasture Co.","Brand","No","Conventional Dairy","USA (Wisconsin)","D",2.79),
    ("sourcream-sb-organic","Dairy & Eggs","Sour Cream","EveryDay Basics Organic Sour Cream (16 oz)","EveryDay Basics","Off-Brand","Yes","Pasture-Raised","USA (Vermont)","D",3.99),
    ("sourcream-brand-organic","Dairy & Eggs","Sour Cream","Green Pasture Co. Organic Sour Cream (16 oz)","Green Pasture Co.","Brand","Yes","Pasture-Raised","USA (Vermont)","D",4.99),
    # Meat & Seafood (additions)
    ("pork-sb-factory","Meat & Seafood","Pork Chops","EveryDay Basics Pork Chops (1 lb)","EveryDay Basics","Off-Brand","No","Factory-Farmed","USA (Iowa)","B",4.49),
    ("pork-brand-factory","Meat & Seafood","Pork Chops","Heritage Range Pork Chops (1 lb)","Heritage Range","Brand","No","Factory-Farmed","USA (Iowa)","B",5.99),
    ("pork-sb-freerange","Meat & Seafood","Pork Chops","EveryDay Basics Free-Range Pork Chops (1 lb)","EveryDay Basics","Off-Brand","No","Free-Range","USA (Iowa)","B",7.49),
    ("pork-brand-freerange-org","Meat & Seafood","Pork Chops","Heritage Range Organic Free-Range Pork Chops (1 lb)","Heritage Range","Brand","Yes","Free-Range","USA (Iowa)","B",9.99),
    # Bakery & Grains (additions)
    ("tortillas-sb-flour","Bakery & Grains","Tortillas","EveryDay Basics Flour Tortillas (10 ct)","EveryDay Basics","Off-Brand","No","Standard","USA","B",2.29),
    ("tortillas-brand-flour","Bakery & Grains","Tortillas","Golden Harvest Flour Tortillas (10 ct)","Golden Harvest","Brand","No","Standard","USA","B",3.29),
    ("tortillas-sb-wheat","Bakery & Grains","Tortillas","EveryDay Basics Whole Wheat Tortillas (10 ct)","EveryDay Basics","Off-Brand","No","Standard","USA","A",2.79),
    ("tortillas-brand-wheat-org","Bakery & Grains","Tortillas","Golden Harvest Organic Whole Wheat Tortillas (10 ct)","Golden Harvest","Brand","Yes","Standard","USA","A",4.49),
    ("oats-sb-rolled","Bakery & Grains","Oats","EveryDay Basics Rolled Oats (18 oz)","EveryDay Basics","Off-Brand","No","Standard","USA","A",2.49),
    ("oats-brand-rolled","Bakery & Grains","Oats","Golden Harvest Rolled Oats (18 oz)","Golden Harvest","Brand","No","Standard","USA","A",3.49),
    ("oats-sb-organic","Bakery & Grains","Oats","EveryDay Basics Organic Rolled Oats (18 oz)","EveryDay Basics","Off-Brand","Yes","Standard","USA","A",3.99),
    ("oats-brand-organic","Bakery & Grains","Oats","Golden Harvest Organic Rolled Oats (18 oz)","Golden Harvest","Brand","Yes","Standard","USA","A",5.49),
    # Pantry & Condiments (additions)
    ("sauce-sb-reg","Pantry & Condiments","Pasta Sauce","EveryDay Basics Marinara Pasta Sauce (24 oz jar)","EveryDay Basics","Off-Brand","No","Standard","USA","B",1.99),
    ("sauce-brand-reg","Pantry & Condiments","Pasta Sauce","Golden Harvest Marinara Pasta Sauce (24 oz jar)","Golden Harvest","Brand","No","Standard","USA","B",2.99),
    ("sauce-sb-organic","Pantry & Condiments","Pasta Sauce","EveryDay Basics Organic Marinara Pasta Sauce (24 oz jar)","EveryDay Basics","Off-Brand","Yes","Standard","USA","B",3.49),
    ("sauce-brand-organic-sm","Pantry & Condiments","Pasta Sauce","Golden Harvest Organic San Marzano Pasta Sauce (24 oz jar)","Golden Harvest","Brand","Yes","Standard","Italy","B",5.49),
    ("honey-sb-std","Pantry & Condiments","Honey","EveryDay Basics Honey (12 oz)","EveryDay Basics","Off-Brand","No","Standard","Argentina","D",3.99),
    ("honey-brand-std","Pantry & Condiments","Honey","Golden Harvest Honey (12 oz)","Golden Harvest","Brand","No","Standard","Argentina","D",5.49),
    ("honey-sb-raw-local","Pantry & Condiments","Honey","EveryDay Basics Raw Local Honey (12 oz)","EveryDay Basics","Off-Brand","No","Raw & Unfiltered","USA (Vermont)","D",6.99),
    ("honey-brand-raw-local-org","Pantry & Condiments","Honey","Golden Harvest Organic Raw Local Honey (12 oz)","Golden Harvest","Brand","Yes","Raw & Unfiltered","USA (Vermont)","D",8.99),
    # Snacks & Beverages (additions)
    ("soda-sb-reg","Snacks & Beverages","Soda","EveryDay Basics Cola (12-pack cans)","EveryDay Basics","Off-Brand","No","Standard","USA","E",4.49),
    ("soda-brand-reg","Snacks & Beverages","Soda","Pinnacle Foods Cola (12-pack cans)","Pinnacle Foods","Brand","No","Standard","USA","E",5.99),
    ("soda-sb-diet","Snacks & Beverages","Soda","EveryDay Basics Diet Cola (12-pack cans)","EveryDay Basics","Off-Brand","No","Standard","USA","C",4.99),
    ("soda-brand-diet","Snacks & Beverages","Soda","Pinnacle Foods Diet Cola (12-pack cans)","Pinnacle Foods","Brand","No","Standard","USA","C",6.49),
    # Plant-Based & Vegetarian (new category)
    ("tofu-sb-std","Plant-Based & Vegetarian","Tofu","EveryDay Basics Firm Tofu (14 oz)","EveryDay Basics","Off-Brand","No","Standard Soybeans","USA (Iowa)","A",1.99),
    ("tofu-brand-std","Plant-Based & Vegetarian","Tofu","Terra Verde Foods Firm Tofu (14 oz)","Terra Verde Foods","Brand","No","Standard Soybeans","USA (Iowa)","A",2.99),
    ("tofu-sb-organic","Plant-Based & Vegetarian","Tofu","EveryDay Basics Organic Non-GMO Firm Tofu (14 oz)","EveryDay Basics","Off-Brand","Yes","Organic Non-GMO Soybeans","USA (Iowa)","A",3.49),
    ("tofu-brand-organic","Plant-Based & Vegetarian","Tofu","Terra Verde Foods Organic Non-GMO Firm Tofu (14 oz)","Terra Verde Foods","Brand","Yes","Organic Non-GMO Soybeans","Canada","A",4.49),
    ("hummus-sb-reg","Plant-Based & Vegetarian","Hummus","EveryDay Basics Classic Hummus (10 oz)","EveryDay Basics","Off-Brand","No","Standard","USA","B",2.99),
    ("hummus-brand-reg","Plant-Based & Vegetarian","Hummus","Terra Verde Foods Classic Hummus (10 oz)","Terra Verde Foods","Brand","No","Standard","USA","B",3.99),
    ("hummus-sb-organic","Plant-Based & Vegetarian","Hummus","EveryDay Basics Organic Classic Hummus (10 oz)","EveryDay Basics","Off-Brand","Yes","Standard","USA","B",4.49),
    ("hummus-brand-organic","Plant-Based & Vegetarian","Hummus","Terra Verde Foods Organic Classic Hummus (10 oz)","Terra Verde Foods","Brand","Yes","Standard","USA","B",5.49),
    # Frozen (new category)
    ("pizza-sb-reg","Frozen","Frozen Pizza","EveryDay Basics Frozen Pepperoni Pizza (12 in)","EveryDay Basics","Off-Brand","No","Standard","USA","D",3.99),
    ("pizza-brand-reg","Frozen","Frozen Pizza","Pinnacle Foods Frozen Pepperoni Pizza (12 in)","Pinnacle Foods","Brand","No","Standard","USA","D",5.49),
    ("pizza-sb-organic","Frozen","Frozen Pizza","EveryDay Basics Organic Frozen Margherita Pizza (12 in)","EveryDay Basics","Off-Brand","Yes","Standard","USA","C",6.49),
    ("pizza-brand-organic","Frozen","Frozen Pizza","Pinnacle Foods Organic Frozen Margherita Pizza (12 in)","Pinnacle Foods","Brand","Yes","Standard","USA","C",7.99),
]

HEADERS = ["Item ID","Category","Product Type","Item Name","Brand","Brand Tier","Organic","Sourcing Practice","Country of Origin","Nutri-Score","Price"]

wb = Workbook()

# ---- Catalog sheet ----
ws = wb.active
ws.title = "Catalog"
ws.append(HEADERS)
for col in range(1, len(HEADERS) + 1):
    c = ws.cell(1, col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

for row in ROWS:
    ws.append(row)

last_row = len(ROWS) + 1
for r in range(2, last_row + 1):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(r, col)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center") if col != 4 else Alignment(horizontal="left", vertical="center")
    price_cell = ws.cell(r, 11)
    price_cell.number_format = '$#,##0.00'
    nutri_cell = ws.cell(r, 10)
    grade = nutri_cell.value
    nutri_cell.fill = PatternFill("solid", start_color=NUTRI_COLORS[grade])
    nutri_cell.font = Font(name=FONT, bold=True, color="FFFFFF" if grade != "C" else "000000", size=10)

widths = [22, 18, 16, 46, 16, 11, 9, 24, 18, 11, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{last_row}"

# ---- Legend sheet ----
legend = wb.create_sheet("Legend")
legend_rows = [
    ("Column", "Meaning"),
    ("Item ID", "Stable machine-readable key for this product variant (will map to the app's catalog data file)."),
    ("Category", "Top-level grocery aisle grouping."),
    ("Product Type", "The specific kind of product (e.g. Apples, Milk) — each type has several variants below."),
    ("Brand Tier", "Brand = a named brand; Off-Brand = the store-brand line, 'EveryDay Basics'. Off-Brand is always cheaper than Brand for an otherwise-equivalent variant."),
    ("Organic", "Whether the product is certified organic."),
    ("Sourcing Practice", "Ethicality-of-sourcing label, category-appropriate: e.g. Factory-Farmed vs Free-Range (meat/eggs), Farmed vs Wild-Caught (seafood), Fair-Trade Certified vs Standard (coffee/chocolate/bananas), Locally Grown vs Imported (domestic produce)."),
    ("Country of Origin", "Where the product is sourced from; part of the ethicality-of-sourcing dimension alongside Sourcing Practice."),
    ("Nutri-Score", "European front-of-pack nutrition grade, A (healthiest) to E (least healthy), color-coded to the official scheme. Reflects the food itself, not its brand/organic/ethics status — e.g. a raw apple is grade A whether organic or not. Varies within a product type only where a genuinely different formulation exists (e.g. plain corn flakes vs. frosted, white vs. whole-wheat bread)."),
    ("Price", "Fictional USD retail price. Within any matched pair, Off-Brand < Brand, and conventional < organic/ethically-sourced."),
]
for r in legend_rows:
    legend.append(r)
legend["A1"].font = HEADER_FONT
legend["B1"].font = HEADER_FONT
legend["A1"].fill = HEADER_FILL
legend["B1"].fill = HEADER_FILL
for row in legend.iter_rows(min_row=2, max_row=legend.max_row):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
legend.column_dimensions["A"].width = 20
legend.column_dimensions["B"].width = 100

# ---- Summary sheet ----
summary = wb.create_sheet("Summary")
summary.append(["Category", "Item Count", "Avg Price (Off-Brand)", "Avg Price (Brand)"])
for col in range(1, 5):
    c = summary.cell(1, col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

categories = sorted(set(r[1] for r in ROWS))
for i, cat in enumerate(categories, start=2):
    summary.cell(i, 1, cat).font = BODY_FONT
    summary.cell(i, 2, f'=COUNTIF(Catalog!B:B,A{i})').font = BODY_FONT
    summary.cell(i, 3, f'=AVERAGEIFS(Catalog!K:K,Catalog!B:B,A{i},Catalog!F:F,"Off-Brand")').font = BODY_FONT
    summary.cell(i, 3).number_format = '$#,##0.00'
    summary.cell(i, 4, f'=AVERAGEIFS(Catalog!K:K,Catalog!B:B,A{i},Catalog!F:F,"Brand")').font = BODY_FONT
    summary.cell(i, 4).number_format = '$#,##0.00'

total_row = len(categories) + 2
summary.cell(total_row, 1, "TOTAL").font = Font(name=FONT, bold=True)
summary.cell(total_row, 2, f"=SUM(B2:B{total_row - 1})").font = Font(name=FONT, bold=True)

summary.column_dimensions["A"].width = 20
summary.column_dimensions["B"].width = 12
summary.column_dimensions["C"].width = 20
summary.column_dimensions["D"].width = 16

import os
wb.save(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "catalog-design.xlsx"))
print(f"Wrote {len(ROWS)} items")
